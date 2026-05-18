"""Populate `data2/SAEM_Hackathon_Hackathon_Class_Disposition_Results.xlsx`
with the production-model predictions:

  - Drug Class      (0–4, 5-class with Siren Spark) ← task1 cascade
  - Disposition Class (0=Discharge, 1=Floor, 2=ICU) ← task2 rforest

Inputs:
  derived/phase2/task1_drug_predictions_phase2_5class.csv
  derived/phase2/task2_disposition_predictions_phase2.csv
  data2/SAEM_Hackathon_Hackathon_Class_Disposition_Results.xlsx

Output (in-place):
  data2/SAEM_Hackathon_Hackathon_Class_Disposition_Results.xlsx
  (a backup of the original template is copied to .pre_fill.xlsx
  before any writes)
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent
DATA2 = ROOT / "data2"
DERIVED = ROOT / "derived"
PHASE2 = DERIVED / "phase2"

DELIVERABLE = DATA2 / "SAEM_Hackathon_Hackathon_Class_Disposition_Results.xlsx"
TASK1_PRED = PHASE2 / "task1_drug_predictions_phase2_5class.csv"
TASK2_PRED = PHASE2 / "task2_disposition_predictions_phase2.csv"

CLASS_DRUG = {0: "No Drug", 1: "Kraken Candy", 2: "Triton Tabs",
              3: "Coral Dust", 4: "Siren Spark"}
CLASS_DISPO = {0: "Discharge", 1: "Floor", 2: "ICU"}


def main() -> None:
    # Sanity: every required file exists.
    for p in (DELIVERABLE, TASK1_PRED, TASK2_PRED):
        if not p.exists():
            raise FileNotFoundError(p)

    # Load predictions
    t1 = pd.read_csv(TASK1_PRED)
    t2 = pd.read_csv(TASK2_PRED)
    print(f"Loaded task1: {t1.shape} (cols: {list(t1.columns)[:5]}...)")
    print(f"Loaded task2: {t2.shape} (cols: {list(t2.columns)[:5]}...)")

    # Backup template before writing
    backup = DELIVERABLE.with_suffix(".pre_fill.xlsx")
    shutil.copy2(DELIVERABLE, backup)
    print(f"Backed up template -> {backup.name}")

    # Open the workbook with openpyxl to preserve formatting/header text
    wb = openpyxl.load_workbook(DELIVERABLE)
    ws = wb.active
    print(f"Worksheet: {ws.title}  dims: "
          f"{ws.max_row} rows x {ws.max_column} cols")

    # Map encounter_id → row index. Row 1 is header; data starts row 2.
    header = [c.value for c in ws[1]]
    print(f"Header: {header}")

    # Build lookup tables from predictions
    t1_map = dict(zip(t1["encounter_id"].astype(str),
                       t1["drug_class"].astype(int)))
    t2_map = dict(zip(t2["encounter_id"].astype(str),
                       t2["disposition_class"].astype(int)))

    filled = {"drug": 0, "dispo": 0}
    missing_drug = []
    missing_dispo = []
    drug_counts = {i: 0 for i in CLASS_DRUG}
    dispo_counts = {i: 0 for i in CLASS_DISPO}

    for r in range(2, ws.max_row + 1):
        eid = str(ws.cell(row=r, column=1).value).strip()
        if not eid or eid == "None":
            continue
        if eid in t1_map:
            cls = int(t1_map[eid])
            ws.cell(row=r, column=2, value=cls)
            filled["drug"] += 1
            drug_counts[cls] += 1
        else:
            missing_drug.append(eid)
        if eid in t2_map:
            cls = int(t2_map[eid])
            ws.cell(row=r, column=3, value=cls)
            filled["dispo"] += 1
            dispo_counts[cls] += 1
        else:
            missing_dispo.append(eid)

    wb.save(DELIVERABLE)
    print(f"\nFilled {filled['drug']} drug-class cells, "
          f"{filled['dispo']} disposition-class cells")
    if missing_drug:
        print(f"  MISSING drug predictions for: {missing_drug[:5]}"
              f"{' ...' if len(missing_drug) > 5 else ''}")
    if missing_dispo:
        print(f"  MISSING dispo predictions for: {missing_dispo[:5]}"
              f"{' ...' if len(missing_dispo) > 5 else ''}")

    print("\nDrug-class distribution:")
    for k, name in CLASS_DRUG.items():
        print(f"  {k} = {name:14s}: {drug_counts[k]:>4d}")
    print("\nDisposition-class distribution:")
    for k, name in CLASS_DISPO.items():
        print(f"  {k} = {name:9s}: {dispo_counts[k]:>4d}")

    # Verify round-trip
    verify = pd.read_excel(DELIVERABLE, sheet_name=0, engine="openpyxl")
    n_filled_drug = int(verify.iloc[:, 1].notna().sum())
    n_filled_dispo = int(verify.iloc[:, 2].notna().sum())
    print(f"\nRound-trip verify: drug filled = {n_filled_drug}, "
          f"dispo filled = {n_filled_dispo} (of {len(verify)} rows)")
    print(f"\nWrote: {DELIVERABLE}")


if __name__ == "__main__":
    main()
